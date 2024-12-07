# -*- coding: utf-8 -*-
"""
Created on Mon Mar  13 17:22:56 2023

@author: Albert dos Santos
VERSÃO FINAL
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from NotaQualis import nota_qualis
from xml.dom import minidom
import pandas as pd
import os
from read_pdf_table import read_pdf_table
import xml.etree.ElementTree as et

# # Armazena o nome dos curriculos em xml
# curriculos = []
# #Lista e armazena os arquivos xml presentes na pasta curriculos
# for f in glob.glob('curriculos/permanentes/*.xml'):    
# 	#Adiciona o nome de cada arquivo xml ao final da lista curriculos[]
# 	curriculos.append(f)   

# Inicializa a lista de currículos
curriculos = []

# Caminho para a pasta onde os arquivos XML estão localizados
caminho_pasta = 'curriculos/todos'

# Lista e armazena os arquivos XML presentes na pasta curriculos
for f in os.listdir(caminho_pasta):
    if f.endswith('.xml'):
        # Adiciona o nome de cada arquivo XML ao final da lista curriculos[]
        curriculos.append(os.path.join(caminho_pasta, f))

# Linha e coluna da planilha de artigos por professor
linha = 2
coluna = 1
# Anos que serão analisados. Para alterar basta adicionar ou retirar o ano da lista abaixo
anos_analisados = [2021, 2022, 2023, 2024]
# anos_analisados = [2020, 2021, 2022, 2023, 2024]


# Cria a planilha que terá o resultado o final da análise
planilha_de_resultado = Workbook()
# Obtema aba padrão que apresentará os artigos publicados por cada curriculo/professor na planilha principal
aba_artigos_por_professor = planilha_de_resultado.active
# Renomeia a aba
aba_artigos_por_professor.title = u'Artigos_por_Professor'
# Inserindo o nome das colunas da planilha
aba_artigos_por_professor.cell(1, 1, u'Professor')
aba_artigos_por_professor.cell(1, 2, u'Tipo')
aba_artigos_por_professor.cell(1, 3, u'Titulo')
aba_artigos_por_professor.cell(1, 4, u'Ano')
aba_artigos_por_professor.cell(1, 5, u'DOI')
aba_artigos_por_professor.cell(1, 6, u'Titulo Periodico ou Revista')
aba_artigos_por_professor.cell(1, 7, u'Autores')
aba_artigos_por_professor.cell(1, 8, u'Estrato')
aba_artigos_por_professor.cell(1, 9, u'Nota')
aba_artigos_por_professor.cell(1,10, u'Aderente')

# Lê a planilha com a base eventos de computação utlizando o pandas (FUNÇÃO)
base_de_eventos = pd.read_excel('QualisEventosComp.xlsx')

# Obtem a quantidade total de eventos e a quantidade de colunas que descrevem estes eventos
total_de_eventos = base_de_eventos.shape[0]
colunas_de_eventos = base_de_eventos.shape[1]
# O Dataframe pandas é convertido para um array numpy. Isso facilitará o acesso aos dados
base_de_eventos = base_de_eventos.to_numpy()
# print(base_de_eventos[1][7])

# Lê a planilha com a base de periódicos utlizando o pandas (FUNÇÃO)
base_de_periodicos = pd.read_excel('Periodicos.xlsx')
# Obtem a quantidade total de eventos e a quantidade de colunas que descrevem estes eventos
total_de_periodicos = base_de_periodicos.shape[0]
colunas_de_periodicos = base_de_periodicos.shape[1]
# O Dataframe pandas é convertido para um array numpy. Isso facilitará o acesso aos dados
base_de_periodicos = base_de_periodicos.to_numpy()
# print(base_de_periodicos[0][1])


# Leitura do pdf com os artigos aderentes
df_artigos_aderentes = read_pdf_table('periodicos_aderentes_das_engenharias_iv_-2017-2020.pdf')

"""
    Percorre cada um dos curriculos
"""
for nome_do_curriculo in curriculos:

    """
        Obtem os artigos publicados em eventos
    """
	# Lê o arquivo xml (FUNÇÃO)
    # curriculo_xml = open(nome_do_curriculo)
    # curriculo = minidom.parse(curriculo_xml)

    try:
        with open(nome_do_curriculo, 'r', encoding='latin-1') as file:
            content = file.read()
            curriculo = minidom.parseString(content)
            # Processa o conteúdo do XML conforme necessário
    except Exception as e:
        print(f"Erro ao processar {nome_do_curriculo}: {e}")

    #Obtém o nome do professor
    dados_gerais_do_professor = curriculo.getElementsByTagName('DADOS-GERAIS')
    for dados in dados_gerais_do_professor:
        nome_do_professor = dados.attributes['NOME-COMPLETO'].value

    # Tipo de publicação que será analisada
    tipo = 'Evento'
    # Lista que armazenará os autores de cada trabalho
    autores_do_trabalho_em_evento = ''
    estrato_do_evento = 'SEM QUALIS'
    nota_evento = 0

    # Obtem todos os trabalhos publicados em eventos
    trabalhos_em_eventos = curriculo.getElementsByTagName('TRABALHO-EM-EVENTOS')
    # Analisa cada trabalho publicado em eventos
    for trabalho_em_eventos in trabalhos_em_eventos:
        aba_artigos_por_professor.cell(linha, 1, nome_do_professor)
        # Obtem as informações do trabalho publicado em evento
        # Como estamos trabalhando com xml, as informações estão nos nós filhos
        trabalho_em_eventos = trabalho_em_eventos.childNodes
        # Analisa cada informação sobre o trabalho contida nos nós filhos
        for dados_do_trabalho_em_evento in trabalho_em_eventos:
            if dados_do_trabalho_em_evento.localName == 'DADOS-BASICOS-DO-TRABALHO':
                print(nome_do_professor)

                print(tipo)
                aba_artigos_por_professor.cell(linha, 2, tipo)

                titulo_do_trabalho_em_evento = dados_do_trabalho_em_evento.attributes['TITULO-DO-TRABALHO'].value
                print(titulo_do_trabalho_em_evento)
                aba_artigos_por_professor.cell(linha, 3, titulo_do_trabalho_em_evento)

                ano_do_trabalho_em_evento = dados_do_trabalho_em_evento.attributes['ANO-DO-TRABALHO'].value
                print(ano_do_trabalho_em_evento)
                aba_artigos_por_professor.cell(linha, 4, ano_do_trabalho_em_evento)


                doi_trabalho_em_evento = dados_do_trabalho_em_evento.attributes['DOI'].value
                print(doi_trabalho_em_evento)
                aba_artigos_por_professor.cell(linha, 5, doi_trabalho_em_evento)


            elif dados_do_trabalho_em_evento.localName == 'DETALHAMENTO-DO-TRABALHO':
                nome_do_evento = dados_do_trabalho_em_evento.attributes['TITULO-DOS-ANAIS-OU-PROCEEDINGS'].value
                print(nome_do_evento)
                aba_artigos_por_professor.cell(linha, 6, nome_do_evento)

            elif dados_do_trabalho_em_evento.localName == 'AUTORES':
                # Une os autores em uma unica string separando-os por ponto e vírgula
                autores_do_trabalho_em_evento = autores_do_trabalho_em_evento + dados_do_trabalho_em_evento.attributes['NOME-COMPLETO-DO-AUTOR'].value + ';'
        
        print(autores_do_trabalho_em_evento)
        aba_artigos_por_professor.cell(linha, 7, autores_do_trabalho_em_evento)
        autores_do_trabalho_em_evento = ''

        # Adiciona o estrato do evento
        # Começa pela linha 1 pois estamos desconsiderando o cabeçalho
        for linha1 in range(1, total_de_eventos): 
            # Obtem o nome do evento
            evento = base_de_eventos[linha1][1]
            if evento == nome_do_evento:
                estrato_do_evento = base_de_eventos[linha1][7]
                print(estrato_do_evento)
                aba_artigos_por_professor.cell(linha, 8, estrato_do_evento)
                break
            
            
        if estrato_do_evento == 'SEM QUALIS':
            aba_artigos_por_professor.cell(linha, 8, estrato_do_evento)

        # Obtem a nota de acordo com o estrado
        nota_evento = nota_qualis(tipo, estrato_do_evento)
        print(nota_evento)
        aba_artigos_por_professor.cell(linha, 9, nota_evento)

        # A coluna aderente refere-se apenas a periódicos. Portanto, não se aplica a eventos
        aba_artigos_por_professor.cell(linha, 10, 'Não')


        estrato_do_evento = 'SEM QUALIS'
        nota_evento = 0
        linha = linha + 1












    """
        Obtem os artigos publicados em periódicos
    """
    # Tipo de publicação que será analisada
    tipo = 'Periodico'
    # Lista que armazenará os autores de cada trabalho
    autores_do_trabalho_em_periodico = ''
    estrato_do_periodico = 'SEM QUALIS'
    nota_periodico = 0

    # Obtem todos os trabalhos publicados em periódicos
    trabalhos_em_periodicos = curriculo.getElementsByTagName('ARTIGO-PUBLICADO')
    # Analisa cada trabalho publicado em periodicos
    for trabalho_em_periodicos in trabalhos_em_periodicos:
        aba_artigos_por_professor.cell(linha, 1, nome_do_professor)
        # Obtem as informações do trabalho publicado
        # Como estamos trabalhando com xml, as informações estão nos nós filhos
        trabalho_em_periodicos = trabalho_em_periodicos.childNodes
        # Analisa cada informação sobre o trabalho contida nos nós filhos
        for dados_do_trabalho_em_periodico in trabalho_em_periodicos:
            if dados_do_trabalho_em_periodico.localName == 'DADOS-BASICOS-DO-ARTIGO':
                
                print(tipo)
                aba_artigos_por_professor.cell(linha, 2, tipo)

                titulo_do_trabalho_em_periodico = dados_do_trabalho_em_periodico.attributes['TITULO-DO-ARTIGO'].value
                print(titulo_do_trabalho_em_periodico)
                aba_artigos_por_professor.cell(linha, 3, titulo_do_trabalho_em_periodico)


                ano_do_trabalho_em_periodico = dados_do_trabalho_em_periodico.attributes['ANO-DO-ARTIGO'].value
                print(ano_do_trabalho_em_periodico)
                aba_artigos_por_professor.cell(linha, 4, ano_do_trabalho_em_periodico)


                doi_trabalho_em_periodico = dados_do_trabalho_em_periodico.attributes['DOI'].value
                print(doi_trabalho_em_periodico)
                aba_artigos_por_professor.cell(linha, 5, doi_trabalho_em_periodico)


            elif dados_do_trabalho_em_periodico.localName == 'DETALHAMENTO-DO-ARTIGO':
                nome_do_periodico = dados_do_trabalho_em_periodico.attributes['TITULO-DO-PERIODICO-OU-REVISTA'].value
                print(nome_do_periodico)
                aba_artigos_por_professor.cell(linha, 6, nome_do_periodico)


                issn_do_periodico_publicado = dados_do_trabalho_em_periodico.attributes['ISSN'].value
                issn_do_periodico_publicado_para_checar_aderentes = issn_do_periodico_publicado
                issn_do_periodico_publicado = issn_do_periodico_publicado[:-1]

                # Verifica se o ISSN do periódico publicado está na lista de periódicos aderentes  
                print(f'ISSN {issn_do_periodico_publicado_para_checar_aderentes}, nome: {nome_do_periodico}')
                # print(df_artigos_aderentes['ISSN'].values)
                if issn_do_periodico_publicado_para_checar_aderentes in df_artigos_aderentes['ISSN'].values:
                    aba_artigos_por_professor.cell(linha, 10, 'Sim')
                else:
                    aba_artigos_por_professor.cell(linha, 10, 'Não')

            elif dados_do_trabalho_em_periodico.localName == 'AUTORES':
                # Une os autores 
                autores_do_trabalho_em_periodico = autores_do_trabalho_em_periodico + dados_do_trabalho_em_periodico.attributes['NOME-COMPLETO-DO-AUTOR'].value + ';'
        
        print(autores_do_trabalho_em_periodico)
        aba_artigos_por_professor.cell(linha, 7, autores_do_trabalho_em_periodico)

        autores_do_trabalho_em_periodico = ''

        # Adiciona o estrato do periodico
        # Começa pela linha 1 pois estamos desconsiderando o cabeçalho
        for linha1 in range(1, total_de_periodicos): 
            # Obtem o issn do periódico
            issn = base_de_periodicos[linha1][0]
            # Retira o "-" do ISSN
            auxiliar = issn.split("-")
            issn = ''.join(auxiliar)

            issn = issn[:-1]
            
            # Não inclui o último número na comparação pois, na base, alguns períodicos possuem um X no lugar do último número
            if issn == issn_do_periodico_publicado:
                estrato_do_periodico = base_de_periodicos[linha1][2]
                print(estrato_do_periodico)
                aba_artigos_por_professor.cell(linha, 8, estrato_do_periodico)
                break


        if estrato_do_periodico == 'SEM QUALIS':
            aba_artigos_por_professor.cell(linha, 8, estrato_do_periodico)

        # Obtem a nota de acordo com o estrado
        nota_periodico = nota_qualis(tipo, estrato_do_periodico)
        print(nota_periodico)
        aba_artigos_por_professor.cell(linha, 9, nota_periodico)

        estrato_do_periodico = 'SEM QUALIS'
        nota_periodico = 0
        linha = linha + 1
    
#salva todos os artigos obtidos em arquivo xls
planilha_de_resultado.save('Lattes_extrator.xlsx')




"""
    Filtra para os anos analisados
    Para filtrar é mais fácil carregar a planilha em Dataframe e realizar a filtragem utilizando o pandas
"""
# Lê a planilha salva e transforma para dataframe
df_planilha_de_resultado = pd.read_excel('Lattes_extrator.xlsx')
# Na coluna ano substiti a string "onic" por "2024"
df_planilha_de_resultado['Ano'] = df_planilha_de_resultado['Ano'].replace('onic', '2024')
# Tranforma a coluna Ano em inteiro
df_planilha_de_resultado['Ano'] = df_planilha_de_resultado['Ano'].astype(int)
# Filtra de acordo com os anos presentes na variável "anos_analisados". "Ano" representa a coluna Ano do dataframe
df_planilha_de_resultado = df_planilha_de_resultado.query("Ano in @anos_analisados")
# Salva o dataframe filtrado em formato de planilha. Obs: sem os indíces
df_planilha_de_resultado.to_excel('Lattes_extrator.xlsx', index=False)






"""
    Corrige as notas
"""
# Lê a planilha com os artigos por professor utlizando o pandas (FUNÇÃO)
artigos_dos_professores = pd.read_excel('Lattes_extrator.xlsx')
# Revome duplicatas
artigos_dos_professores = artigos_dos_professores.drop_duplicates()
# Verifica se há linhas com o mesmo valor de Professor e Título e remove uma delas
artigos_dos_professores = artigos_dos_professores.drop_duplicates(subset=['Professor', 'Titulo'], keep='first')
# Obtem a lista de artigos publicados de todos os curriculos analisados
lista_de_artigos_publicados = artigos_dos_professores['Titulo'].to_list()
# Obtem a quantidade total de artigos e a quantidade de colunas que descrevem estes artigos
total_de_artigos_publicados = artigos_dos_professores.shape[0]
colunas_de_artigos_publicados = artigos_dos_professores.shape[1]
# O Dataframe pandas é convertido para um array numpy. Isso facilitará o acesso aos dados
artigos_dos_professores = artigos_dos_professores.to_numpy()
# print(artigos_dos_professores[1][2])

for indice_artigo in range(total_de_artigos_publicados):
    titulo_do_artigo_analisado = artigos_dos_professores[indice_artigo][2]
    cont = (str(lista_de_artigos_publicados).upper()).count(str(titulo_do_artigo_analisado).upper())
    if cont > 1:
        nota_atual = artigos_dos_professores[indice_artigo][8]
        nota_corrigida = nota_atual / cont
        artigos_dos_professores[indice_artigo][8] = nota_corrigida


nome_das_colunas = ['Professor','Tipo','Titulo','Ano', 'DOI', 'Titulo Periodico ou Revista', 'Autores', 'Estrato','Nota', 'Aderente']
df_artigos_dos_professores = pd.DataFrame(artigos_dos_professores, columns=nome_das_colunas)
df_artigos_dos_professores.to_excel('Lattes_extrator.xlsx', sheet_name='Artigos_por_professor', index=False)






"""
    Cria aba com nota dos professores por ano
"""
# Carrega a planilha principal
planilha_de_resultado = load_workbook(filename = 'Lattes_extrator.xlsx')

try:
    # Tenta abrir a aba na planilha
    aba_notas_por_ano = planilha_de_resultado['Notas_por_ano']
except:
    # Cria nova aba na planilha principal
    aba_notas_por_ano = planilha_de_resultado.create_sheet(u'Notas_por_ano')  

# Limpa a aba para receber os novos dados 
aba_notas_por_ano.delete_rows(2, aba_notas_por_ano.max_row-1)

# Inserindo o nome das colunas da planilha
aba_notas_por_ano.cell(1, 1, u'Professor')

coluna = 2
# Adiciona as colunas para cada ano que será analisado
for ano in anos_analisados:
    aba_notas_por_ano.cell(1, coluna, u'{}'.format(ano))
    coluna = coluna + 1


linha = 2
coluna = 1
# Lê a planilha principal e transforma para dataframe
df_planilha_de_resultado = pd.read_excel('Lattes_extrator.xlsx')
# Obtem o nome de cada professor
lista_de_professores = df_planilha_de_resultado['Professor'].unique()

# Percorre as informações de cada um dos professores
for professor in lista_de_professores:

    # Adiciona o nome do professor na planilha
    aba_notas_por_ano.cell(linha, coluna, professor)
    coluna = coluna + 1

    for ano in anos_analisados:
        # O filtro gera os artigos do professor "professor" publicados no ano "ano"
        # Ou seja, teremos as publicações do professor em cada ano
        artigos_publicados_em_um_ano = df_planilha_de_resultado.query("Ano == @ano and Professor == @professor")
        # Obtem a soma de todos os artigos publicados pelo professor no ano em questão
        nota_total_do_ano = artigos_publicados_em_um_ano['Nota'].sum()

        # Adiciona a nota do ano
        aba_notas_por_ano.cell(linha, coluna, nota_total_do_ano)
        coluna = coluna + 1
    
    linha = linha + 1
    coluna = 1


planilha_de_resultado.save('Lattes_extrator.xlsx')









"""
    Cria aba com os artigos publicados e os autores
"""
# Carrega a planilha principal
planilha_de_resultado = load_workbook(filename = 'Lattes_extrator.xlsx')

try:
    # Tenta abrir a aba na planilha
    aba_artigo_e_autores = planilha_de_resultado['Artigos_e_autores']
except:
    # Cria nova aba na planilha principal
    aba_artigo_e_autores = planilha_de_resultado.create_sheet(u'Artigos_e_autores')  

# Limpa a aba para receber os novos dados 
aba_artigo_e_autores.delete_rows(2, aba_artigo_e_autores.max_row-1)

# Inserindo o nome das colunas da planilha
aba_artigo_e_autores.cell(1, 1, u'Titulo')
aba_artigo_e_autores.cell(1, 2, u'Ano')
aba_artigo_e_autores.cell(1, 3, u'Periodico/Revista')
aba_artigo_e_autores.cell(1, 4, u'Estrato')
aba_artigo_e_autores.cell(1, 5, u'Nota')
aba_artigo_e_autores.cell(1, 6, u'Autor')

# Lê a planilha principal e transforma para dataframe
df_planilha_de_resultado = pd.read_excel('Lattes_extrator.xlsx')

# Filtra para obter somente os periódicos
df_planilha_de_resultado = df_planilha_de_resultado.query('Tipo == "Periodico"')

# Obtem o nome dos artigos publicados em periódicos
lista_de_artigos_publicados_em_periodico = df_planilha_de_resultado['Titulo'].unique()

linha = 2
# Percorre os artigos publicados em periódicos
for artigo in lista_de_artigos_publicados_em_periodico:
    # Obtem os professores que publicaram o artigo
    df_professores_autores = df_planilha_de_resultado.query('Titulo == @artigo')
    # Quantidade de professores que são autores do artigo
    quantidade_de_professores_autores = df_professores_autores.shape[0]

    # Convertemos para numpy array para facilitar o acesso aos dados
    professores_autores = df_professores_autores.to_numpy()

    # Adiciona o nome do artigo
    aba_artigo_e_autores.cell(linha, 1, artigo)
    # Adiciona o ano do artigo
    aba_artigo_e_autores.cell(linha, 2, professores_autores[0][3])
    # Adiciona o periodico
    aba_artigo_e_autores.cell(linha, 3, professores_autores[0][5])
    # Adiciona o estrato
    aba_artigo_e_autores.cell(linha, 4, professores_autores[0][7])
    # Adiciona a nota
    aba_artigo_e_autores.cell(linha, 5, professores_autores[0][8])
    
    # Se for mais de um, é necessário adicionar todos
    if quantidade_de_professores_autores > 1:
        for indice_do_autor in range(quantidade_de_professores_autores):
            # Nome de um autor
            nome = professores_autores[indice_do_autor][0]

            # Obtem o valor da cabeçalho na planilha referente aos autores
            valor_da_cabecalho_autor = aba_artigo_e_autores.cell(1, indice_do_autor + 6)
            # Se ela estiver vazia devemos criar o nome para cabeçalho dos autores na planilha
            if not valor_da_cabecalho_autor.value:
                #criar o titulo da coluna na planilha
                aba_artigo_e_autores.cell(1, indice_do_autor + 6, u'Autor {}'.format(indice_do_autor+1))
            
            # Adiciona o nome do autor na planilha
            aba_artigo_e_autores.cell(linha, indice_do_autor + 6, nome)

    else:
        # Nome de um autor
        nome = professores_autores[0][0]
        # Adiciona o nome do autor na planilha
        aba_artigo_e_autores.cell(linha, 6, nome)

    linha = linha + 1


planilha_de_resultado.save('Lattes_extrator.xlsx')


"""
    Integração das Patentes
    Agora vamos adicionar as duas abas: Patentes e pontuacao_patentes.
    Caso hajam patentes duplicadas (mesma patente para mais de um professor), a nota será dividida igualmente, assim como foi feito para os artigos.
"""

dados_patentes = []
for nome_do_curriculo in curriculos:
    try:
        tree = et.parse(nome_do_curriculo)
        xml_root = tree.getroot()
    except Exception as e:
        print(f"Erro ao processar {nome_do_curriculo}: {e}")
        continue

    dados_gerais = xml_root.find('DADOS-GERAIS')
    nome_docente = dados_gerais.attrib.get('NOME-COMPLETO', '') if dados_gerais is not None else "Nome não encontrado"
    
    for patente in xml_root.iter('PATENTE'):
        dados_basicos = patente.find('DADOS-BASICOS-DA-PATENTE')
        detalhamento = patente.find('DETALHAMENTO-DA-PATENTE')
        
        if dados_basicos is not None and detalhamento is not None:
            ano_desenvolvimento = dados_basicos.attrib.get("ANO-DESENVOLVIMENTO", "")
            titulo_patente = dados_basicos.attrib.get("TITULO", "")
            
            if ano_desenvolvimento in [str(a) for a in anos_analisados]:
                historico = detalhamento.find('HISTORICO-SITUACOES-PATENTE')
                descricao_situacao = historico.attrib.get("DESCRICAO-SITUACAO-PATENTE", "").strip() if historico is not None else ""
                
                if descricao_situacao == "Concessão":
                    situacao_final = "Concessão"
                elif descricao_situacao.startswith("Depósito"):
                    situacao_final = "Depósito"
                else:
                    situacao_final = "Indefinido"
                
                dados_patentes.append({
                    "Professor": nome_docente,
                    "Nome da Patente": titulo_patente,
                    "Ano": ano_desenvolvimento,
                    "Status": situacao_final
                })

df_patentes = pd.DataFrame(dados_patentes)
with pd.ExcelWriter('Lattes_extrator.xlsx', engine='openpyxl', mode='a') as writer:
    if df_patentes.empty:
        # Cria abas vazias
        df_patentes_vazio = pd.DataFrame(columns=["Professor", "Nome da Patente", "Ano", "Status", "Nota"])
        df_patentes_vazio.to_excel(writer, sheet_name='Patentes', index=False)
        df_pontuacao_vazio = pd.DataFrame(columns=["Professor", "2021", "2022", "2023", "2024"])
        df_pontuacao_vazio.to_excel(writer, sheet_name='pontuacao_patentes', index=False)
    else:
        df_patentes = df_patentes[["Professor", "Nome da Patente", "Ano", "Status"]]

        def calcular_pontuacao(status):
            if status == "Depósito":
                return nota_qualis("PERIODICO", "A2")  # 0.875
            elif status == "Concessão":
                return 2 * nota_qualis("PERIODICO", "A1") # 2 * 1 = 2
            else:
                return 0

        df_pont = df_patentes.copy()
        df_pont['Pontuacao'] = df_pont['Status'].apply(calcular_pontuacao)

        # Ajuste por duplicatas (divisão da pontuação)
        lista_de_patentes_publicadas = df_pont['Nome da Patente'].to_list()
        patentes_array = df_pont.to_numpy()
        total_de_patentes = df_pont.shape[0]

        for i in range(total_de_patentes):
            titulo_patente = patentes_array[i][1]  
            cont = (str(lista_de_patentes_publicadas).upper()).count(str(titulo_patente).upper())
            if cont > 1:
                nota_atual = patentes_array[i][4]  
                nota_corrigida = nota_atual / cont
                patentes_array[i][4] = nota_corrigida

        df_pont = pd.DataFrame(patentes_array, columns=["Professor", "Nome da Patente", "Ano", "Status", "Pontuacao"])
        # Renomeando Pontuacao para Nota
        df_pont.rename(columns={"Pontuacao": "Nota"}, inplace=True)

        # Agora df_pont possui a coluna Nota corrigida
        # Pivot para pontuacao_patentes
        df_pivot = df_pont.pivot_table(index='Professor', columns='Ano', values='Nota', aggfunc='sum', fill_value=0)

        for year in anos_analisados:
            year_str = str(year)
            if year_str not in df_pivot.columns:
                df_pivot[year_str] = 0

        df_pivot = df_pivot[[str(y) for y in anos_analisados]].reset_index()

        df_pont.to_excel(writer, sheet_name='Patentes', index=False)
        df_pivot.to_excel(writer, sheet_name='pontuacao_patentes', index=False)

print("Processo concluído :) ")
